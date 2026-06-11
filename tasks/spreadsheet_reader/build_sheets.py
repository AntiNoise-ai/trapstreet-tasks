"""Source-of-truth generator for the spreadsheet_reader task.

Builds a real .xlsx per case with an authentic trap baked into the file
(text-typed numbers, accounting parentheses, hidden rows, a stale SUM formula,
a thousands-unit header, duplicate keys). The gold answer is computed in Python
from the same source values, so it can't drift from the file.

Run from this directory:

    python3 build_sheets.py

Each case emits:
  inputs/<id>/data.xlsx
  inputs/<id>/question.txt
  expected/<id>/answer.json   (leading_numeric matcher, like pdf_reader)
and the aggregate traptask.yaml + gold.cases.json.

The trap in every case: the *naive* read gives a tidy-but-wrong number; the
correct answer needs you to notice the mess (units, cell type, hidden rows,
a wrong formula, accounting signs, duplicates).
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent


def _write(wb: Workbook, cid: str) -> None:
    d = HERE / "inputs" / cid
    d.mkdir(parents=True, exist_ok=True)
    wb.save(d / "data.xlsx")


def _question(cid: str, text: str) -> None:
    (HERE / "inputs" / cid).mkdir(parents=True, exist_ok=True)
    (HERE / "inputs" / cid / "question.txt").write_text(text.strip() + "\n")


def _answer(cid: str, value: float, category: str, difficulty: str,
            decoy: float, note: str, tolerance: float = 0.01) -> dict:
    (HERE / "expected" / cid).mkdir(parents=True, exist_ok=True)
    ans = {
        "id": cid,
        "answer": str(int(value)) if value == int(value) else str(value),
        "type": "numeric",
        "matchers": [{"kind": "leading_numeric", "value": float(value), "tolerance": tolerance}],
        "category": category,
        "difficulty": difficulty,
        "_decoy_naive_answer": decoy,
        "_notes": note,
    }
    (HERE / "expected" / cid / "answer.json").write_text(json.dumps(ans, indent=2) + "\n")
    return ans


CASES_META = []


def case_units_in_header():
    cid = "units_in_header"
    rows = [("North", 120.5), ("South", 89.0), ("East", 42.25), ("West", 300.0)]
    wb = Workbook(); ws = wb.active; ws.title = "Q3"
    ws.append(["Region", "Revenue ($ thousands)"])
    for r in rows:
        ws.append(list(r))
    _write(wb, cid)
    gold = round(sum(v for _, v in rows) * 1000, 2)      # 551750.0
    decoy = round(sum(v for _, v in rows), 2)             # 551.75
    _question(cid, """
The attached spreadsheet (data.xlsx) reports quarterly revenue by region.
Read the column header carefully.

What is the TOTAL revenue across all regions, expressed in US dollars (not thousands)?
Answer with a single number and nothing else.
""")
    CASES_META.append((cid, "units", "easy",
        _answer(cid, gold, "units", "easy", decoy,
                "Header says '$ thousands'. Naive sum = 551.75; correct = x1000 = 551750.")))


def case_numbers_as_text():
    cid = "numbers_as_text"
    # Stored as TEXT strings with thousands separators -> SUM()/float() trip up.
    vals = ["1,234.50", "2,000", "750.25", "3,499.99", "12,000"]
    wb = Workbook(); ws = wb.active; ws.title = "Ledger"
    ws.append(["Item", "Amount"])
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=i + 1, column=1, value=f"item-{i}")
        cell = ws.cell(row=i + 1, column=2, value=v)
        cell.number_format = "@"  # force text type
    _write(wb, cid)
    gold = round(sum(float(v.replace(",", "")) for v in vals), 2)  # 19484.74
    _question(cid, """
The attached spreadsheet (data.xlsx) has an 'Amount' column.

What is the total of the 'Amount' column? Answer with a single number and nothing else.
""")
    CASES_META.append((cid, "types", "medium",
        _answer(cid, gold, "types", "medium", 0.0,
                "Amounts are TEXT cells with comma separators. SUM() returns 0 / float() raises; "
                "correct total = 19484.74.")))


def case_parentheses_negatives():
    cid = "parentheses_negatives"
    # Accounting convention: parens = negative. Stored as text.
    vals = ["4,500", "(1,200)", "2,300", "(800)", "950"]
    wb = Workbook(); ws = wb.active; ws.title = "GL"
    ws.append(["Entry", "Amount"])
    for i, v in enumerate(vals, 1):
        ws.cell(row=i + 1, column=1, value=f"entry-{i}")
        cell = ws.cell(row=i + 1, column=2, value=v)
        cell.number_format = "@"
    _write(wb, cid)

    def parse(v: str) -> float:
        v = v.replace(",", "")
        if v.startswith("(") and v.endswith(")"):
            return -float(v[1:-1])
        return float(v)

    gold = round(sum(parse(v) for v in vals), 2)                 # 5750.0
    decoy = round(sum(abs(parse(v)) for v in vals), 2)           # 9750.0 (signs ignored)
    _question(cid, """
The attached spreadsheet (data.xlsx) is a general-ledger extract. It uses
accounting format: amounts shown in parentheses are NEGATIVE.

What is the NET total of the 'Amount' column? Answer with a single number and nothing else.
""")
    CASES_META.append((cid, "signs", "medium",
        _answer(cid, gold, "signs", "medium", decoy,
                "Parentheses = negative. Treating them as positive gives 9750; correct net = 5750.")))


def case_total_row_double_count():
    cid = "total_row_double_count"
    amounts = [100, 200, 150, 300, 250, 175, 225, 400, 125]  # line items
    wb = Workbook(); ws = wb.active; ws.title = "Invoices"
    ws.append(["Line", "Amount"])
    for i, a in enumerate(amounts, 1):
        ws.cell(row=i + 1, column=1, value=f"L{i}")
        ws.cell(row=i + 1, column=2, value=a)
    # A literal TOTAL row sitting IN the Amount column -> a naive column .sum() doubles it.
    ws.cell(row=len(amounts) + 2, column=1, value="TOTAL")
    ws.cell(row=len(amounts) + 2, column=2, value=sum(amounts))
    _write(wb, cid)
    gold = sum(amounts)              # 1925
    decoy = sum(amounts) * 2         # 3850 (column .sum() incl. the TOTAL row)
    _question(cid, """
The attached spreadsheet (data.xlsx) lists invoice line items in the 'Amount'
column and ends with a 'TOTAL' row.

What is the sum of the LINE ITEMS only (exclude the TOTAL row)? Answer with a
single number and nothing else.
""")
    CASES_META.append((cid, "total_row", "hard",
        _answer(cid, gold, "total_row", "hard", decoy,
                "A TOTAL row lives in the Amount column; a naive df['Amount'].sum() double-counts "
                "to 3850. Correct line-item sum = 1925.")))


def case_wrong_sheet():
    cid = "wrong_sheet"
    wb = Workbook()
    # First (default) sheet = a DRAFT summary -> what pd.read_excel() reads by default.
    s1 = wb.active; s1.title = "Summary"
    s1.append(["Note", "Amount"])
    for label, amt in [("draft-a", 2000), ("draft-b", 2000), ("draft-c", 2000)]:
        s1.append([label, amt])   # naive first-sheet sum = 6000 (the decoy)
    # Real data lives on a second sheet named in the question.
    s2 = wb.create_sheet("FY2026_Actuals")
    s2.append(["Quarter", "Revenue"])
    actuals = [("Q1", 1000), ("Q2", 2000), ("Q3", 1500), ("Q4", 3000)]
    for q, r in actuals:
        s2.append([q, r])
    _write(wb, cid)
    gold = sum(r for _, r in actuals)   # 7500
    decoy = 6000                        # first-sheet draft total
    _question(cid, """
The attached workbook (data.xlsx) has more than one sheet.

What is the TOTAL revenue on the sheet named 'FY2026_Actuals'? Answer with a
single number and nothing else.
""")
    CASES_META.append((cid, "multi_sheet", "hard",
        _answer(cid, gold, "multi_sheet", "hard", decoy,
                "Real data is on the 'FY2026_Actuals' sheet (sum 7500). pd.read_excel() defaults to "
                "the first sheet ('Summary', sum 6000) — the decoy.")))


def case_dedup_messy():
    cid = "dedup_messy_customers"
    # Same customers, but differing only by case / surrounding whitespace.
    customers = ["Acme", "Beta", "acme", "Gamma", "BETA ", "Acme ", " gamma "]
    wb = Workbook(); ws = wb.active; ws.title = "Orders"
    ws.append(["OrderID", "Customer"])
    for i, c in enumerate(customers, 1):
        ws.cell(row=i + 1, column=1, value=1000 + i)
        ws.cell(row=i + 1, column=2, value=c)
    _write(wb, cid)
    gold = len({c.strip().casefold() for c in customers})   # 3 (acme/beta/gamma)
    decoy = len(set(customers))                              # 7 (exact-distinct)
    _question(cid, """
The attached spreadsheet (data.xlsx) lists orders with a 'Customer' column. Some
customer names differ only in letter case or surrounding spaces.

How many UNIQUE customers are there, treating names that differ only in case or
surrounding whitespace as the SAME customer? Answer with a single number and nothing else.
""")
    CASES_META.append((cid, "dedup", "medium",
        _answer(cid, gold, "dedup", "medium", decoy,
                "7 exact-distinct strings but only 3 real customers (Acme/Beta/Gamma) once case and "
                "whitespace are normalized. A naive nunique() = 7.")))


def main():
    case_units_in_header()
    case_numbers_as_text()
    case_parentheses_negatives()
    case_total_row_double_count()
    case_wrong_sheet()
    case_dedup_messy()

    # traptask.yaml
    lines = ["dirs:", "  inputs: inputs/", "  expected: expected/", "", "cases:"]
    gold = []
    for cid, cat, diff, ans in CASES_META:
        desc = ans["_notes"].replace('"', "'")
        lines += [f"- id: {cid}", f'  description: "{desc}"', "  tags:",
                  f"  - {cat}", f"  - {diff}", ""]
        gold.append({"id": cid, "category": cat, "difficulty": diff,
                     "answer": ans["answer"], "decoy_naive_answer": ans["_decoy_naive_answer"],
                     "notes": ans["_notes"]})
    lines += ["judge:", "  cmd: python3 judge.py", "", "grader:", "  cmd: python3 grader.py", ""]
    (HERE / "traptask.yaml").write_text("\n".join(lines))
    (HERE / "gold.cases.json").write_text(json.dumps(gold, indent=2) + "\n")
    print(f"generated {len(CASES_META)} spreadsheet cases into {HERE}")


if __name__ == "__main__":
    main()
